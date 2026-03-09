"""Excel 匯出功能（任務清單 + 依負責人工作日誌分頁）"""

from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

EXPORT_COLUMNS = [
    ('編號',       'id',              8),
    ('專案',       'project_name',   20),
    ('標題',       'title',          35),
    ('狀態',       'status',         12),
    ('優先級',     'priority',       10),
    ('類別',       'category',       25),
    ('負責人',     'assignee',       15),
    ('到期日',     'due_date',       14),
    ('開始日期',   'start_date',     14),
    ('預估週數',   'estimated_weeks', 10),
    ('標籤',       'tags',           20),
    ('描述',       'description',    50),
    ('建立時間',   'created_at',     20),
    ('更新時間',   'updated_at',     20),
]

WORKLOG_COLUMNS = [
    ('任務標題', 'task_title', 30),
    ('負責人',   'assignee',   15),
    ('日期',     'log_date',   14),
    ('工時(天)', 'hours',      10),
    ('內容',     'content',    50),
]

MILESTONE_COLUMNS = [
    ('專案',     'project_name', 20),
    ('名稱',     'name',         30),
    ('目標日期', 'target_date',  14),
    ('描述',     'description',  40),
]

FONT_NAME = 'Microsoft JhengHei UI'


def export_tasks_to_excel(filepath, tasks, project_name="全部專案",
                          project_lookup=None, task_tags_lookup=None,
                          db=None, milestones=None):
    """匯出任務至 Excel。

    結構：
        Sheet 1 - 任務清單（全部任務）
        Sheet per Assignee - 該負責人的工作日誌
    """
    if project_lookup is None:
        project_lookup = {}
    if task_tags_lookup is None:
        task_tags_lookup = {}

    wb = Workbook()

    # ── 共用樣式 ──
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4',
                              fill_type='solid')
    header_font = Font(name=FONT_NAME, size=10, bold=True, color='FFFFFF')
    body_font = Font(name=FONT_NAME, size=10)
    thin_border = Border(bottom=Side(style='thin', color='999999'))

    priority_fills = {
        '緊急': PatternFill(start_color='FFC7CE', end_color='FFC7CE',
                          fill_type='solid'),
        '高':   PatternFill(start_color='FFE699', end_color='FFE699',
                          fill_type='solid'),
    }
    status_fills = {
        '已完成': PatternFill(start_color='C6EFCE', end_color='C6EFCE',
                            fill_type='solid'),
        '進行中': PatternFill(start_color='BDD7EE', end_color='BDD7EE',
                            fill_type='solid'),
    }

    # ══════════════════════════════════════════════════════
    # Sheet 1: 全部任務
    # ══════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "任務清單"
    _write_task_sheet(ws, tasks, project_name, project_lookup,
                      task_tags_lookup, header_fill, header_font,
                      body_font, thin_border, priority_fills, status_fills)

    # ══════════════════════════════════════════════════════
    # Sheet per Assignee: 依負責人分頁 — 顯示工作日誌
    # ══════════════════════════════════════════════════════
    if db:
        # 按負責人收集工作日誌
        assignee_logs = {}  # {assignee: [(task_title, log), ...]}
        for task in tasks:
            logs = db.get_work_logs(task.id)
            if logs:
                key = task.assignee if task.assignee else '未指派'
                if key not in assignee_logs:
                    assignee_logs[key] = []
                for log in logs:
                    assignee_logs[key].append((task.title, task.assignee, log))

        for assignee_name in sorted(assignee_logs.keys()):
            safe_name = _safe_sheet_name(assignee_name)
            ws_a = wb.create_sheet(title=safe_name)
            _write_assignee_worklog_sheet(
                ws_a, assignee_name, assignee_logs[assignee_name],
                project_name, header_fill, header_font, body_font,
                thin_border)

    # ══════════════════════════════════════════════════════
    # Sheet: 里程碑
    # ══════════════════════════════════════════════════════
    if milestones:
        ws_ms = wb.create_sheet(title="里程碑")
        _write_milestone_sheet(ws_ms, milestones,
                               project_lookup or {},
                               header_fill, header_font,
                               body_font, thin_border)

    wb.save(filepath)


def _safe_sheet_name(name):
    """處理 Excel 工作表名稱限制"""
    safe = name[:28]
    for ch in ['/', '\\', '[', ']', '*', '?', ':']:
        safe = safe.replace(ch, '_')
    return safe


def _write_task_sheet(ws, tasks, title_text, project_lookup, task_tags_lookup,
                      header_fill, header_font, body_font, thin_border,
                      priority_fills, status_fills):
    """寫入一頁任務表"""
    num_cols = len(EXPORT_COLUMNS)

    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1, end_column=num_cols)
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = (f"專案: {title_text}  —  "
                        f"匯出時間: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    title_cell.font = Font(name=FONT_NAME, size=14, bold=True)
    title_cell.alignment = Alignment(horizontal='left')

    for col_idx, (header, _, width) in enumerate(EXPORT_COLUMNS, 1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    for row_idx, task in enumerate(tasks, 4):
        for col_idx, (_, field, _) in enumerate(EXPORT_COLUMNS, 1):
            if field == 'project_name':
                value = project_lookup.get(task.project_id, '')
            elif field == 'tags':
                value = ', '.join(task_tags_lookup.get(task.id, []))
            elif field == 'estimated_weeks':
                value = task.estimated_weeks if task.estimated_weeks else ''
            else:
                value = getattr(task, field, '')

            cell = ws.cell(row=row_idx, column=col_idx, value=value or '')
            cell.font = body_font
            cell.alignment = Alignment(
                vertical='center',
                wrap_text=(field == 'description'))
            cell.border = thin_border

            if field == 'priority' and task.priority in priority_fills:
                cell.fill = priority_fills[task.priority]
            if field == 'status' and task.status in status_fills:
                cell.fill = status_fills[task.status]

    if tasks:
        last_row = 3 + len(tasks)
        ws.auto_filter.ref = f"A3:{get_column_letter(num_cols)}{last_row}"
    ws.freeze_panes = 'A4'


def _write_assignee_worklog_sheet(ws, assignee_name, log_entries,
                                   project_name, header_fill, header_font,
                                   body_font, thin_border):
    """寫入負責人工作日誌分頁"""
    num_cols = len(WORKLOG_COLUMNS)

    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1, end_column=num_cols)
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = (f"工作日誌 - {assignee_name} ({project_name})  —  "
                        f"匯出時間: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    title_cell.font = Font(name=FONT_NAME, size=14, bold=True)
    title_cell.alignment = Alignment(horizontal='left')

    for col_idx, (header, _, width) in enumerate(WORKLOG_COLUMNS, 1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # 按日期排序
    sorted_entries = sorted(log_entries, key=lambda x: x[2].log_date,
                             reverse=True)

    row_idx = 4
    for task_title, assignee, log in sorted_entries:
        data = {
            'task_title': task_title,
            'assignee': assignee or '',
            'log_date': log.log_date,
            'hours': log.hours,
            'content': log.content,
        }
        for col_idx, (_, field, _) in enumerate(WORKLOG_COLUMNS, 1):
            value = data.get(field, '')
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = body_font
            cell.alignment = Alignment(
                vertical='center',
                wrap_text=(field == 'content'))
            cell.border = thin_border
        row_idx += 1

    if row_idx > 4:
        ws.auto_filter.ref = f"A3:{get_column_letter(num_cols)}{row_idx - 1}"
    ws.freeze_panes = 'A4'


def _write_milestone_sheet(ws, milestones, project_lookup,
                            header_fill, header_font, body_font, thin_border):
    """寫入里程碑分頁，target_date 轉為 Excel 日期格式"""
    from datetime import date as date_type
    num_cols = len(MILESTONE_COLUMNS)

    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1, end_column=num_cols)
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = (f"里程碑清單  —  "
                        f"匯出時間: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    title_cell.font = Font(name=FONT_NAME, size=14, bold=True)
    title_cell.alignment = Alignment(horizontal='left')

    for col_idx, (header, _, width) in enumerate(MILESTONE_COLUMNS, 1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    sorted_ms = sorted(milestones, key=lambda m: m.target_date or '')
    for row_idx, ms in enumerate(sorted_ms, 4):
        for col_idx, (_, field, _) in enumerate(MILESTONE_COLUMNS, 1):
            if field == 'project_name':
                value = project_lookup.get(ms.project_id, '')
            elif field == 'target_date':
                try:
                    value = datetime.strptime(ms.target_date, '%Y-%m-%d').date()
                except (ValueError, TypeError):
                    value = ms.target_date or ''
            else:
                value = getattr(ms, field, '') or ''

            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = body_font
            cell.alignment = Alignment(vertical='center',
                                       wrap_text=(field == 'description'))
            cell.border = thin_border
            if field == 'target_date' and isinstance(value, date_type):
                cell.number_format = 'YYYY-MM-DD'

    if len(sorted_ms) > 0:
        ws.auto_filter.ref = f"A3:{get_column_letter(num_cols)}{3 + len(sorted_ms)}"
    ws.freeze_panes = 'A4'
