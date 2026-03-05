"""Excel 匯入邏輯：讀取 .xlsx 並驗證後匯入資料庫"""

import re
from datetime import datetime

# 表頭對照（Excel 欄位名 → Task 屬性）
HEADER_MAP = {
    '標題': 'title',
    '狀態': 'status',
    '優先級': 'priority',
    '類別': 'category',
    '負責人': 'assignee',
    '到期日': 'due_date',
    '描述': 'description',
    '開始日期': 'start_date',
    '預估週數': 'estimated_weeks',
    '標籤': 'tags',
}

VALID_STATUSES = {'待辦', '進行中', '審核中', '已完成'}
VALID_PRIORITIES = {'緊急', '高', '中', '低'}


def import_tasks_from_excel(filepath, db, project_id):
    """匯入 Excel 至指定專案。回傳 (imported_count, errors_list)。
    採 all-or-nothing 策略：有任何錯誤則完全不匯入。
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        return 0, ["缺少 openpyxl 套件，請執行: pip install openpyxl"]

    try:
        wb = load_workbook(filepath, read_only=True, data_only=True)
    except Exception as e:
        return 0, [f"無法開啟檔案: {e}"]

    ws = wb.active

    # 掃描前 10 列，找含「標題」的表頭列
    header_row = None
    header_map = {}   # col_idx -> field_name
    for row_idx in range(1, 11):
        for col_idx, cell in enumerate(ws[row_idx], 1):
            val = str(cell.value).strip() if cell.value else ''
            if val == '標題':
                header_row = row_idx
                break
        if header_row:
            break

    if header_row is None:
        wb.close()
        return 0, ["找不到表頭列（需包含「標題」欄位）"]

    # 建立欄位對照
    for col_idx, cell in enumerate(ws[header_row], 1):
        val = str(cell.value).strip() if cell.value else ''
        if val in HEADER_MAP:
            header_map[col_idx] = HEADER_MAP[val]

    if 'title' not in header_map.values():
        wb.close()
        return 0, ["表頭列缺少必要的「標題」欄位"]

    # 解析每列資料
    errors = []
    records = []

    data_start = header_row + 1
    for row_idx, row in enumerate(ws.iter_rows(min_row=data_start), start=data_start):
        record = {}
        for col_idx, cell in enumerate(row, 1):
            field = header_map.get(col_idx)
            if field:
                val = cell.value
                if val is not None:
                    if isinstance(val, datetime):
                        record[field] = val.strftime('%Y-%m-%d')
                    else:
                        record[field] = str(val).strip()
                else:
                    record[field] = ''

        # 跳過空列
        if not record.get('title'):
            continue

        # 驗證
        title = record.get('title', '').strip()
        if not title:
            errors.append(f"第 {row_idx} 列: 標題不可為空")
            continue

        # 狀態驗證
        status = record.get('status', '待辦').strip()
        if status and status not in VALID_STATUSES:
            errors.append(
                f"第 {row_idx} 列, 欄位「狀態」: "
                f"「{status}」不是有效狀態 (待辦/進行中/審核中/已完成)")

        # 優先級驗證
        priority = record.get('priority', '中').strip()
        if priority and priority not in VALID_PRIORITIES:
            errors.append(
                f"第 {row_idx} 列, 欄位「優先級」: "
                f"「{priority}」不是有效優先級 (緊急/高/中/低)")

        # 日期格式驗證
        for date_field, label in [('due_date', '到期日'), ('start_date', '開始日期')]:
            d = record.get(date_field, '').strip()
            if d:
                if not re.match(r'^\d{4}-\d{2}-\d{2}$', d):
                    errors.append(
                        f"第 {row_idx} 列, 欄位「{label}」: "
                        f"「{d}」格式應為 YYYY-MM-DD")

        # 預估週數驗證
        ew = record.get('estimated_weeks', '').strip()
        if ew:
            try:
                ew_int = int(ew)
                if ew_int < 0:
                    errors.append(
                        f"第 {row_idx} 列, 欄位「預估週數」: 不可為負數")
                record['estimated_weeks'] = ew_int
            except ValueError:
                errors.append(
                    f"第 {row_idx} 列, 欄位「預估週數」: 「{ew}」必須為整數")
        else:
            record['estimated_weeks'] = None

        record['title'] = title
        record['status'] = status or '待辦'
        record['priority'] = priority or '中'
        records.append(record)

    wb.close()

    # all-or-nothing
    if errors:
        return 0, errors

    if not records:
        return 0, ["檔案中無可匯入的資料"]

    # 匯入
    count = 0
    for rec in records:
        tags_str = rec.get('tags', '')
        tags = [t.strip() for t in tags_str.split(',') if t.strip()] if tags_str else None

        db.create_task(
            project_id=project_id,
            title=rec['title'],
            description=rec.get('description', ''),
            status=rec['status'],
            priority=rec['priority'],
            category=rec.get('category', ''),
            assignee=rec.get('assignee', ''),
            due_date=rec.get('due_date') or None,
            start_date=rec.get('start_date') or None,
            estimated_weeks=rec.get('estimated_weeks'),
            tags=tags,
        )
        count += 1

    return count, []
