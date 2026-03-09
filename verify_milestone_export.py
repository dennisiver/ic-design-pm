"""驗收腳本：驗證匯出的 Excel 含有里程碑分頁且日期格式正確。

使用方式：
  1. 開啟程式，確認至少有一筆 milestone（target_date = 2026-03-20）
  2. 執行匯出，儲存為 test_export.xlsx（放在此腳本同目錄）
  3. 執行：python verify_milestone_export.py test_export.xlsx
"""

import sys
from datetime import date
from openpyxl import load_workbook


def verify(filepath):
    wb = load_workbook(filepath)

    # 1. 里程碑 sheet 存在
    assert '里程碑' in wb.sheetnames, \
        f"FAIL: 找不到「里程碑」分頁，現有分頁：{wb.sheetnames}"
    print("PASS [1] 「里程碑」分頁存在")

    ws = wb['里程碑']

    # 2. 標頭列正確（第3列）
    headers = [ws.cell(3, c).value for c in range(1, 5)]
    assert '目標日期' in headers, \
        f"FAIL: 第3列標頭不含「目標日期」，實際：{headers}"
    date_col = headers.index('目標日期') + 1
    print(f"PASS [2] 標頭正確，「目標日期」在第 {date_col} 欄")

    # 3. 找第一筆資料列，驗證日期格式
    date_cell = None
    for row in ws.iter_rows(min_row=4):
        if row[0].value:  # 專案名稱不為空
            date_cell = row[date_col - 1]
            break

    assert date_cell is not None, "FAIL: 里程碑分頁沒有任何資料列"
    assert isinstance(date_cell.value, date), \
        f"FAIL: 日期欄型別應為 datetime.date，實際為 {type(date_cell.value)!r}，值={date_cell.value!r}"
    print(f"PASS [3] 日期欄型別正確（datetime.date），值={date_cell.value}")

    assert date_cell.number_format == 'YYYY-MM-DD', \
        f"FAIL: number_format 應為 'YYYY-MM-DD'，實際為 {date_cell.number_format!r}"
    print(f"PASS [4] number_format 正確：{date_cell.number_format}")

    # 4. 如果有 2026-03-20 這筆，額外驗證值
    expected = date(2026, 3, 20)
    found = any(
        row[date_col - 1].value == expected
        for row in ws.iter_rows(min_row=4)
        if row[0].value
    )
    if found:
        print(f"PASS [5] 找到 target_date = {expected} 的資料列")
    else:
        print(f"INFO [5] 未找到 {expected}（若測試資料不同可忽略）")

    print("\n全部驗收通過。")


if __name__ == '__main__':
    path = sys.argv[1] if len(sys.argv) > 1 else 'test_export.xlsx'
    verify(path)
